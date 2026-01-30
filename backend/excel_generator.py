from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
from io import BytesIO


class ExcelGenerator:
    # Excel report generator for detection results

    def __init__(self):
        self.colors = {
            'primary': '2D3748',      # Dark gray/blue
            'secondary': '4A5568',    # Medium gray
            'accent': '718096',       # Light gray
            'light': 'EDF2F7',        # Light background
            'white': 'FFFFFF',        # White
            'text': '1A202C'          # Dark text
        }
        
        self.fonts = {
            'title': Font(size=14, bold=True, color=self.colors['primary']),
            'subtitle': Font(size=12, bold=True, color=self.colors['secondary']),
            'header': Font(size=11, bold=True, color=self.colors['white']),
            'body': Font(size=10, color=self.colors['text']),
            'bold': Font(size=10, bold=True, color=self.colors['text']),
            'small': Font(size=9, color=self.colors['accent'])
        }
        
        self.thin_border = Border(
            left=Side(style='thin', color=self.colors['accent']),
            right=Side(style='thin', color=self.colors['accent']),
            top=Side(style='thin', color=self.colors['accent']),
            bottom=Side(style='thin', color=self.colors['accent'])
        )
        
        self.header_fill = PatternFill(
            start_color=self.colors['primary'],
            end_color=self.colors['primary'],
            fill_type='solid'
        )
        
        self.light_fill = PatternFill(
            start_color=self.colors['light'],
            end_color=self.colors['light'],
            fill_type='solid'
        )
    
    def _apply_header_style(self, ws, row, start_col, num_cols):
        for col in range(start_col, start_col + num_cols):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.fonts['header']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
    
    def _apply_table_style(self, ws, start_row, end_row, start_col, num_cols):
        for row in range(start_row, end_row + 1):
            row_fill = PatternFill(
                start_color=self.colors['white'] if row % 2 == 0 else self.colors['light'],
                end_color=self.colors['white'] if row % 2 == 0 else self.colors['light'],
                fill_type='solid'
            )
            
            for col in range(start_col, start_col + num_cols):
                cell = ws.cell(row=row, column=col)
                cell.fill = row_fill
                cell.border = self.thin_border
                cell.font = self.fonts['body']
                
                if col > start_col:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(vertical='center')
    
    def _add_spacer(self, ws, row, height=5):
        ws.row_dimensions[row].height = height
    
    def _add_title(self, ws, cell_ref, title, font_type='title'):
        cell = ws[cell_ref]
        cell.value = title
        cell.font = self.fonts[font_type]
    
    def _add_metadata(self, ws, cell_ref, text):
        cell = ws[cell_ref]
        cell.value = text
        cell.font = self.fonts['small']
        cell.alignment = Alignment(horizontal='center')
    
    def generate_full_report(self, history_entries):
        wb = Workbook()
        ws = wb.active
        ws.title = "Detection Report"
        
        start_col = 2
        ws.merge_cells(f'B1:H1')
        self._add_title(ws, 'B1', "Animal Detection Report")
        ws.merge_cells(f'B2:H2')
        self._add_metadata(ws, 'B2', f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        
        self._add_spacer(ws, 3)
        
        self._add_title(ws, f'B4', "Summary", 'subtitle')
        
        total_entries = len(history_entries)
        total_animals = sum(entry.total_animals for entry in history_entries)
        total_cats = sum(entry.cats_count for entry in history_entries)
        total_dogs = sum(entry.dogs_count for entry in history_entries)
        avg_animals = total_animals / total_entries if total_entries > 0 else 0
        
        stats_data = [
            ["Total Images", total_entries],
            ["Total Animals", total_animals],
            ["Total Cats", total_cats],
            ["Total Dogs", total_dogs],
            ["Average per Image", round(avg_animals, 1)]
        ]
        
        # Write summary statistics starting from column B
        stats_start_row = 6
        for i, (label, value) in enumerate(stats_data):
            ws.cell(row=stats_start_row + i, column=start_col).value = label
            ws.cell(row=stats_start_row + i, column=start_col + 1).value = value
            ws.cell(row=stats_start_row + i, column=start_col).font = self.fonts['bold']
            ws.cell(row=stats_start_row + i, column=start_col + 1).alignment = Alignment(horizontal='center')
        
        for row in range(stats_start_row, stats_start_row + len(stats_data)):
            for col in range(start_col, start_col + 2):
                ws.cell(row=row, column=col).border = self.thin_border
        
        self._add_spacer(ws, stats_start_row + len(stats_data))
        

        history_start_row = stats_start_row + len(stats_data) + 2
        self._add_title(ws, f'B{history_start_row}', "Detection History", 'subtitle')
        
        headers = [
            "ID", "Filename", "Model", 
            "Date", "Time", 
            "Animals", "Cats", "Dogs",
            "Time (s)"
        ]
        
        header_row = history_start_row + 2
        for col, header in enumerate(headers, start_col):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
        
        self._apply_header_style(ws, header_row, start_col, len(headers))
        
        data_start_row = header_row + 1
        for i, entry in enumerate(history_entries):
            date_str = entry.upload_time.strftime('%Y-%m-%d') if entry.upload_time else ""
            time_str = entry.upload_time.strftime('%H:%M') if entry.upload_time else ""
            
            row_data = [
                entry.id,
                entry.filename,
                entry.model_name or "Unknown",
                date_str,
                time_str,
                entry.total_animals,
                entry.cats_count,
                entry.dogs_count,
                round(entry.processing_time, 1) if entry.processing_time else ""
            ]
            
            for j, value in enumerate(row_data, start_col):
                ws.cell(row=data_start_row + i, column=j).value = value
        
        # Apply styling to data table
        if history_entries:
            self._apply_table_style(ws, data_start_row, 
                                   data_start_row + len(history_entries) - 1, 
                                   start_col, len(headers))
        
        column_widths = {
            'B': 8,   # ID
            'C': 25,  # Filename
            'D': 15,  # Model
            'E': 10,  # Date
            'F': 8,   # Time
            'G': 8,   # Animals
            'H': 6,   # Cats
            'I': 6,   # Dogs
            'J': 8    # Time (s)
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        ws2 = wb.create_sheet(title="Model Analysis")
        ws2.merge_cells('B1:G1')
        self._add_title(ws2, 'B1', "Model Performance")
        ws2.merge_cells('B2:G2')
        self._add_metadata(ws2, 'B2', f"Based on {total_entries} detections")
        
        summary_headers = [
            "Model", "Images", 
            "Total Animals", "Cats", "Dogs",
            "Avg Time (s)"
        ]
        
        header_row2 = 4
        for col, header in enumerate(summary_headers, start_col):
            ws2.cell(row=header_row2, column=col).value = header
        
        self._apply_header_style(ws2, header_row2, start_col, len(summary_headers))
        
        model_stats = {}
        for entry in history_entries:
            model_name = entry.model_name or 'Unknown'
            if model_name not in model_stats:
                model_stats[model_name] = {
                    'count': 0,
                    'total_animals': 0,
                    'total_cats': 0,
                    'total_dogs': 0,
                    'total_time': 0
                }
            
            stats = model_stats[model_name]
            stats['count'] += 1
            stats['total_animals'] += entry.total_animals
            stats['total_cats'] += entry.cats_count
            stats['total_dogs'] += entry.dogs_count
            stats['total_time'] += entry.processing_time or 0
        
        data_start_row2 = header_row2 + 1
        row_idx = 0
        for model_name, stats in model_stats.items():
            avg_time = stats['total_time'] / stats['count'] if stats['count'] > 0 else 0
            
            ws2.cell(row=data_start_row2 + row_idx, column=start_col).value = model_name
            ws2.cell(row=data_start_row2 + row_idx, column=start_col + 1).value = stats['count']
            ws2.cell(row=data_start_row2 + row_idx, column=start_col + 2).value = stats['total_animals']
            ws2.cell(row=data_start_row2 + row_idx, column=start_col + 3).value = stats['total_cats']
            ws2.cell(row=data_start_row2 + row_idx, column=start_col + 4).value = stats['total_dogs']
            ws2.cell(row=data_start_row2 + row_idx, column=start_col + 5).value = round(avg_time, 1)
            
            row_idx += 1
        
        if model_stats:
            self._apply_table_style(ws2, data_start_row2, 
                                   data_start_row2 + len(model_stats) - 1, 
                                   start_col, len(summary_headers))
        
        model_widths = {
            'B': 20,   # Model
            'C': 10,   # Images
            'D': 12,   # Total Animals
            'E': 8,    # Cats
            'F': 8,    # Dogs
            'G': 12    # Avg Time
        }
        
        for col_letter, width in model_widths.items():
            ws2.column_dimensions[col_letter].width = width
        
        ws2.row_dimensions[data_start_row2 + len(model_stats) + 2].height = 5
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def generate_single_report(self, entry):
        wb = Workbook()
        ws = wb.active
        ws.title = "Detection Details"
        
        start_col = 2
        
        ws.merge_cells(f'B1:F1')
        self._add_title(ws, 'B1', f"Detection #{entry.id}")
        
        ws.merge_cells(f'B2:F2')
        self._add_metadata(ws, 'B2', f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        
        self._add_spacer(ws, 3)
        
        self._add_title(ws, f'B4', "Image Information", 'subtitle')
        
        # Image information data
        info_data = [
            ["Filename:", entry.filename],
            ["Model:", entry.model_name or "Unknown"],
            ["Date:", entry.upload_time.strftime('%Y-%m-%d') if entry.upload_time else ""],
            ["Time:", entry.upload_time.strftime('%H:%M:%S') if entry.upload_time else ""],
            ["Processing Time:", f"{entry.processing_time or 0:.1f} s"],
            ["Confidence:", f"{(entry.confidence or 0.25) * 100:.0f}%"]
        ]
        
        # Write image information
        info_start_row = 6
        for i, (label, value) in enumerate(info_data):
            ws.cell(row=info_start_row + i, column=start_col).value = label
            ws.cell(row=info_start_row + i, column=start_col + 1).value = value
            ws.cell(row=info_start_row + i, column=start_col).font = self.fonts['bold']
            ws.cell(row=info_start_row + i, column=start_col + 1).alignment = Alignment(horizontal='left')
            
            for col in range(start_col, start_col + 2):
                cell = ws.cell(row=info_start_row + i, column=col)
                cell.border = self.thin_border
                if i % 2 == 0:
                    cell.fill = self.light_fill
        
        self._add_spacer(ws, info_start_row + len(info_data))
        
        summary_start_row = info_start_row + len(info_data) + 2
        self._add_title(ws, f'B{summary_start_row}', "Detection Summary", 'subtitle')
        
        # Summary data
        summary_data = [
            ["Total Animals:", entry.total_animals],
            ["Cats Detected:", entry.cats_count],
            ["Dogs Detected:", entry.dogs_count]
        ]
        
        # Write summary data
        summary_table_start = summary_start_row + 2
        for i, (label, value) in enumerate(summary_data):
            ws.cell(row=summary_table_start + i, column=start_col).value = label
            ws.cell(row=summary_table_start + i, column=start_col + 1).value = value
            ws.cell(row=summary_table_start + i, column=start_col).font = self.fonts['bold']
            ws.cell(row=summary_table_start + i, column=start_col + 1).alignment = Alignment(
                horizontal='center', 
                vertical='center'
            )
            
            for col in range(start_col, start_col + 2):
                cell = ws.cell(row=summary_table_start + i, column=col)
                cell.border = self.thin_border
                if i % 2 == 0:
                    cell.fill = self.light_fill
        
        detections_start = summary_table_start + len(summary_data) + 3
        self._add_title(ws, f'B{detections_start}', "Detected Objects", 'subtitle')
        
        if entry.detections_json:
            detections = json.loads(entry.detections_json)
            
            if detections:
                headers = ["#", "Type", "Confidence", "Area"]
                
                header_row = detections_start + 2
                for col, header in enumerate(headers, start_col):
                    cell = ws.cell(row=header_row, column=col)
                    cell.value = header
                
                self._apply_header_style(ws, header_row, start_col, len(headers))
                
                data_start_row = header_row + 1
                for i, detection in enumerate(detections, 1):
                    confidence_percent = detection.get('confidence', 0) * 100
                    area = detection.get('area', 0)
                    
                    row_data = [
                        i,
                        detection.get('class', 'Unknown').title(),
                        f"{confidence_percent:.1f}%",
                        f"{area:.0f} px²" if area else ""
                    ]
                    
                    for j, value in enumerate(row_data, start_col):
                        ws.cell(row=data_start_row + i - 1, column=j).value = value
                
                self._apply_table_style(ws, data_start_row, 
                                       data_start_row + len(detections) - 1, 
                                       start_col, len(headers))
                
                if detections:
                    stats_start = data_start_row + len(detections) + 2
                    ws.cell(row=stats_start, column=start_col).value = "Detection Statistics:"
                    ws.cell(row=stats_start, column=start_col).font = self.fonts['bold']
                    
                    confidences = [d.get('confidence', 0) for d in detections]
                    avg_confidence = sum(confidences) / len(confidences) * 100 if confidences else 0
                    
                    ws.cell(row=stats_start + 1, column=start_col).value = "Average Confidence:"
                    ws.cell(row=stats_start + 1, column=start_col + 1).value = f"{avg_confidence:.1f}%"
                    ws.cell(row=stats_start + 1, column=start_col + 1).font = self.fonts['bold']
                    
            else:
                ws.cell(row=detections_start + 2, column=start_col).value = "No objects detected in this image"
                ws.cell(row=detections_start + 2, column=start_col).font = self.fonts['body']
        else:
            ws.cell(row=detections_start + 2, column=start_col).value = "No detection data available"
            ws.cell(row=detections_start + 2, column=start_col).font = self.fonts['body']
        
        column_widths = {
            'B': 18,   # Labels
            'C': 22,   # Values
            'D': 12,   # Type/Confidence
            'E': 12,   # Area
            'F': 15    # Extra space
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        last_row = ws.max_row + 2
        ws.cell(row=last_row, column=start_col).value = "Report generated by Animal Detection System"
        ws.cell(row=last_row, column=start_col).font = self.fonts['small']
        ws.cell(row=last_row, column=start_col).alignment = Alignment(horizontal='center')
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer